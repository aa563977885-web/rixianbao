# -*- coding: utf-8 -*-
import json
import os
from datetime import datetime, timedelta

import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSX = os.path.join(ROOT, "outputs", "得物搬砖利润计算器.xlsx")
EXPOSURE = os.path.join(ROOT, "data", "exposure.json")
POOL = os.path.join(ROOT, "data", "pool.json")

EXPOSURE_COL = 12  # L
SIGNAL_COL = 13    # M
HEADER_ROW = 3
FIRST_EMPTY_ROW = 10
LAST_ROW = 200


def load_json(path, default):
    try:
        with open(path, encoding="utf-8") as f:
            return json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        return default


def ratio_for(row_text, pool):
    t = str(row_text or "").strip()
    for item in pool:
        title = str(item.get("title") or "")
        code = str(item.get("article_no") or "")
        if (code and code.lower() in t.lower()) or (title and title in t) or (t and t in title):
            want = item.get("want_count")
            sales = item.get("sales_7d")
            ratio = 0
            if want is not None:
                ratio = round(want / max(int(sales or 0), 1), 2)
            return item.get("carry_id"), ratio
    return None, 0


def signal_formula(row, ratio_txt):
    return f'=IF(L{row}>=5,"避开·全网都知道",IF(AND({ratio_txt}>3,L{row}<3),"可蹲","观察"))'


def main():
    exposure = load_json(EXPOSURE, {})
    pool = load_json(POOL, [])
    wb = openpyxl.load_workbook(XLSX)
    ws = wb["盯款清单"]

    now = datetime.now()
    ws.cell(row=HEADER_ROW, column=EXPOSURE_COL, value="近48h曝光次数")
    ws.cell(row=HEADER_ROW, column=SIGNAL_COL, value="信号建议")

    filled = 0
    for r in range(4, LAST_ROW + 1):
        b = ws.cell(row=r, column=2).value
        if b is not None and str(b).startswith("👇"):
            continue
        is_example = b is not None
        cid, ratio = ratio_for(b, pool) if is_example else (None, 0)
        seen = 0
        if cid:
            info = exposure.get(cid, {})
            last = info.get("last_seen")
            if last:
                try:
                    last_dt = datetime.fromisoformat(last)
                    if last_dt.tzinfo is not None and now.tzinfo is None:
                        last_dt = last_dt.replace(tzinfo=None)
                    if now - last_dt <= timedelta(hours=48):
                        seen = int(info.get("seen_count", 0))
                except ValueError:
                    seen = 0
        ratio_txt = ("%.2f" % ratio) if ratio else "0"
        ws.cell(row=r, column=EXPOSURE_COL, value=seen)
        ws.cell(row=r, column=SIGNAL_COL, value=signal_formula(r, ratio_txt))
        filled += 1

    wb.save(XLSX)
    print(f"盯款清单更新完成：共处理 {filled} 行（含空行预填），列 L/M 已写入 -> {XLSX}")


if __name__ == "__main__":
    main()
