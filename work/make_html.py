# -*- coding: utf-8 -*-
"""
渲染：data/scored.json -> outputs/今日线报.html + outputs/site/index.html
=======================================================================
排序：按稀缺分降序（冷门优选在前，来自 scoring.py 的 verdict 与 score）
卡片：⭐冷门优选/普通徽章 + 想买人数/7日销量/曝光次数/时效标签
缓存：head 三件套 meta + 主链接 ?ts=<unix秒> + 顶部更新时间戳
兼容：用 Excel「今日线报」按 carryId 回填下单教程/涨价预期/校准/说明
"""
import html as H
import json
import os
import re
import sys
import time as _time
from datetime import datetime

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
WORK = os.path.join(ROOT, "work")
BASE = os.path.join(ROOT, "outputs")
XLSX = os.path.join(BASE, "得物搬砖利润计算器.xlsx")
SCORED = os.path.join(ROOT, "data", "scored.json")
OUT = os.path.join(BASE, "今日线报.html")
SITE_OUT = os.path.join(BASE, "site", "index.html")

sys.path.insert(0, WORK)
import scoring  # noqa: E402  复用 age_hours/time_factor

VERDICT_RANK = {"冷门优选": 0, "普通": 1, "黑名单": 2, "过期": 3}


def load_scored():
    try:
        with open(SCORED, encoding="utf-8") as f:
            return json.load(f)
    except FileNotFoundError:
        return None


def load_excel_extra():
    """按 carryId 建 Excel 补充字段索引（教程/涨价预期/校准/说明/发布/码数/销量）。"""
    import openpyxl
    try:
        wb = openpyxl.load_workbook(XLSX)
    except FileNotFoundError:
        return {}
    ws = wb["今日线报"]
    idx = {}
    for r in range(4, ws.max_row + 1):
        link = ws.cell(row=r, column=11).value or ""
        m = re.search(r"carryId=(\d+)", str(link))
        if not m:
            continue
        cid = m.group(1)
        idx[cid] = {
            "publish": ws.cell(row=r, column=1).value,
            "tut": ws.cell(row=r, column=12).value or "",
            "fresh": ws.cell(row=r, column=13).value or "",
            "cal": ws.cell(row=r, column=14).value or "",
            "sizes": ws.cell(row=r, column=15).value or "",
            "sales": ws.cell(row=r, column=16).value or "",
            "expect": ws.cell(row=r, column=17).value or "",
            "note": ws.cell(row=r, column=19).value or "",
        }
    return idx


def fmt(x):
    if x is None:
        return "—"
    return ("%.1f" % x) if isinstance(x, (int, float)) else str(x)


def with_ts(url, ts):
    if not url:
        return url
    sep = "&" if "?" in url else "?"
    return f"{url}{sep}ts={ts}"


def enrich(item, excel, ts):
    cid = str(item.get("carry_id", ""))
    ex = excel.get(cid, {})
    out = dict(item)
    for k in ("tut", "cal", "expect", "note"):
        if k not in out or not out.get(k):
            out[k] = ex.get(k, "")
    if (out.get("published_at") is None) and ex.get("publish"):
        try:
            out["published_at"] = datetime.fromisoformat(str(ex["publish"]).replace(" ", "T")).isoformat(timespec="minutes")
            age = scoring.age_hours(out["published_at"])
            wt, label = scoring.time_factor(age)
            out["time_label"] = label
            out["w_time"] = wt
        except ValueError:
            pass
    out["url"] = with_ts(item.get("url", ""), ts)
    return out


def build_cards(scored, excel, ts):
    cards = []
    for item in scored:
        c = enrich(item, excel, ts)
        net = c.get("net_profit")
        score = c.get("scarcity_score")
        verdict = c.get("verdict") or "普通"
        badges = []
        badges.append(f'<span class="badge cat">{H.escape(str(c.get("category") or "其他"))}</span>')
        if verdict == "冷门优选":
            badges.append('<span class="badge cold">⭐ 冷门优选</span>')
        elif net is not None and net >= 50:
            badges.append('<span class="badge normal">普通</span>')
        badges.append(f'<span class="badge time">{H.escape(str(c.get("time_label") or "时效未知"))}</span>')
        if verdict == "过期":
            badges.append('<span class="badge dead">过期</span>')
        elif verdict == "黑名单":
            badges.append('<span class="badge dead">黑名单</span>')

        rows = []
        rows.append(("<tr><td>货号</td><td>" + H.escape(str(c.get("article_no") or "—")) + "</td></tr>"))
        rows.append(("<tr><td>到手价</td><td>¥" + fmt(c.get("cost")) + "</td></tr>"))
        rows.append(("<tr><td>得物价</td><td>¥" + fmt(c.get("du_price")) + "</td></tr>"))
        rows.append(("<tr><td>净利</td><td>¥" + fmt(net) + "</td></tr>"))
        rows.append(("<tr><td>稀缺分</td><td>" + (str(score) if score is not None else "—") + "</td></tr>"))
        rows.append(("<tr><td>想买人数</td><td>" + (str(c.get("want_count")) if c.get("want_count") is not None else "—") + "</td></tr>"))
        sales = c.get("sales_7d")
        if sales is None:
            sales = (c.get("excel_sales") or "—")
        rows.append(("<tr><td>得物7天销量</td><td>" + H.escape(str(sales)) + "</td></tr>"))
        rows.append(("<tr><td>曝光次数</td><td>" + str(c.get("seen_count") or 0) + " 次</td></tr>"))
        size_txt = c.get("size") or (c.get("excel_sizes") or "")
        color_txt = c.get("color") or ""
        if size_txt and color_txt:
            rows.append(("<tr><td>码数/配色</td><td>" + H.escape(str(color_txt)) + " " + H.escape(str(size_txt)) + "</td></tr>"))
        elif size_txt or color_txt:
            rows.append(("<tr><td>码数/配色</td><td>" + H.escape(str(size_txt or color_txt)) + "</td></tr>"))

        p = ['<div class="card">']
        p.append('<div class="head"><span class="name">🔥 ' + H.escape(str(c.get("title") or "未知品")) + '</span>')
        p.append('<span class="badges">' + "".join(badges) + '</span></div>')
        p.append('<div class="profit">' + verdict_banner(verdict, net) + '</div>')
        p.append('<table class="info">' + "".join(rows) + '</table>')
        if c.get("expect"):
            p.append('<div class="exp">📈 ' + H.escape(str(c["expect"])) + '</div>')
        if c.get("url"):
            p.append('<div class="btn-row">')
            p.append('<a class="btn" href="' + H.escape(c["url"]) + '" target="_blank" rel="noopener">🛒 打开下单页</a>')
            kw = c.get("article_no") or c.get("title") or ""
            p.append('<button class="btn dewu" onclick="copyDewu(this,\'' + H.escape(str(kw)).replace("'", "\\'") + '\')">🔍 得物查价</button>')
            p.append('<button class="btn copy" onclick="copyLink(this,\'' + H.escape(c["url"]).replace("'", "\\'") + '\')">📋 复制链接</button>')
            p.append('</div>')
            p.append('<div class="hint">🔍得物查价：点一下复制货号，去得物App顶部搜索框粘贴即可；📋复制链接：粘到🍑🐶 App 下单更快</div>')
        if c.get("tut"):
            p.append('<details><summary>📖 下单教程</summary><div class="tut">' + H.escape(str(c["tut"])).replace("\n", "<br>") + '</div></details>')
        if c.get("cal"):
            p.append('<div class="cal">📌 校准：' + H.escape(str(c["cal"])) + '</div>')
        if c.get("note"):
            p.append('<div class="note">💬 ' + H.escape(str(c["note"])) + '</div>')
        p.append('</div>')
        cards.append("".join(p))
    return cards


def verdict_banner(verdict, net):
    net_txt = ("+" + fmt(net)) if net is not None else "—"
    if verdict == "冷门优选":
        return '<div class="profit coldline">⭐ 冷门优选　净利 <b>' + net_txt + '</b> 元</div>'
    if verdict == "过期":
        return '<div class="profit redline">⏰ 已过期　净利 <b>' + net_txt + '</b> 元</div>'
    if verdict == "黑名单":
        return '<div class="profit redline">🚫 全网皆知·黑名单　净利 <b>' + net_txt + '</b> 元</div>'
    if net is not None and net >= 50:
        return '<div class="profit greenline">✅ 可做　净利 <b>' + net_txt + '</b> 元</div>'
    return '<div class="profit">⚠️ 核对费率　净利 <b>' + net_txt + '</b> 元</div>'


def main():
    scored = load_scored()
    if not scored:
        print("缺少 data/scored.json，先跑 python work/scoring.py")
        sys.exit(1)
    excel = load_excel_extra()
    ts = int(_time.time())

    def sort_key(item):
        return (VERDICT_RANK.get(item.get("verdict"), 1), -(item.get("scarcity_score") or 0))

    scored = sorted(scored, key=sort_key)
    cards = build_cards(scored, excel, ts)
    cards_html = chr(10).join(cards)

    live_links = []
    for item in scored:
        url = item.get("url")
        if url:
            live_links.append('<a class="chip" href="' + H.escape(with_ts(url, ts)) + '" target="_blank" rel="noopener">🛒 ' + H.escape(str(item.get("title") or "?")) + '</a>')
    live_html = '<div class="chips">' + "".join(live_links) + '</div>'

    n_cold = sum(1 for x in scored if x.get("verdict") == "冷门优选")
    n_ok = sum(1 for x in scored if x.get("net_profit") is not None and x["net_profit"] >= 50)
    n = len(scored)
    stamp = datetime.now().strftime("%Y-%m-%d %H:%M")

    html_doc = TEMPLATE
    html_doc = html_doc.replace("__TS__", str(ts))
    html_doc = html_doc.replace("__STAMP__", stamp)
    html_doc = html_doc.replace("__N__", str(n))
    html_doc = html_doc.replace("__COLD__", str(n_cold))
    html_doc = html_doc.replace("__OK__", str(n_ok))
    html_doc = html_doc.replace("__CARDS__", cards_html)
    html_doc = html_doc.replace("__LIVE_LINKS__", live_html)

    os.makedirs(os.path.dirname(SITE_OUT), exist_ok=True)
    for path in (OUT, SITE_OUT):
        with open(path, "w", encoding="utf-8") as f:
            f.write(html_doc)
        print("OK:", path, os.path.getsize(path), "bytes, cards:", len(cards))


TEMPLATE = """<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<meta http-equiv="Cache-Control" content="no-cache, no-store, must-revalidate">
<meta http-equiv="Pragma" content="no-cache">
<meta http-equiv="Expires" content="0">
<title>得物搬砖 · 差异化线报</title>
<style>
  * { box-sizing: border-box; margin: 0; padding: 0; }
  body { font-family: -apple-system, "PingFang SC", "Microsoft YaHei", sans-serif;
         background: #f2f4f7; color: #1f2328; padding: 16px 12px 40px; }
  .wrap { max-width: 640px; margin: 0 auto; }
  header h1 { font-size: 20px; margin-bottom: 6px; }
  header p { font-size: 12px; color: #666; line-height: 1.7; }
  .tabs { display: flex; gap: 8px; margin: 12px 0; }
  .tab { flex: 1; padding: 10px; border: 1px solid #d9dee5; background: #fff; color: #444;
         border-radius: 10px; font-size: 14px; font-weight: 600; cursor: pointer; }
  .tab.on { background: #1f4e79; border-color: #1f4e79; color: #fff; }
  .stat { display: grid; grid-template-columns: repeat(3, 1fr); gap: 8px; margin-bottom: 14px; }
  .stat > div { background: #fff; border-radius: 10px; padding: 10px; text-align: center; }
  .stat .n { font-size: 20px; font-weight: 700; color: #1f4e79; }
  .stat .t { font-size: 11px; color: #888; }
  .card { background: #fff; border-radius: 10px; padding: 12px; margin-bottom: 12px;
          box-shadow: 0 1px 3px rgba(0,0,0,.06); }
  .head { display: flex; justify-content: space-between; gap: 8px; align-items: flex-start; margin-bottom: 8px; }
  .name { font-size: 15px; font-weight: 700; line-height: 1.4; }
  .badges { display: flex; flex-wrap: wrap; gap: 4px; justify-content: flex-end; max-width: 45%; }
  .badge { font-size: 11px; padding: 2px 6px; border-radius: 6px; white-space: nowrap; }
  .badge.cat { background: #eef1f5; color: #555; }
  .badge.cold { background: #c0392b; color: #fff; font-weight: 700; }
  .badge.normal { background: #e8f0fb; color: #1f4e79; }
  .badge.time { background: #fff4e0; color: #b26a00; }
  .badge.dead { background: #e5e7eb; color: #888; text-decoration: line-through; }
  .profit { font-size: 13px; background: #fdf6ec; color: #8a5a00; border-radius: 8px;
            padding: 8px 10px; margin-bottom: 8px; }
  .profit b { font-size: 17px; }
  .profit.coldline { background: #fdeaea; color: #c0392b; }
  .profit.greenline { background: #e6f6ec; color: #1e7e45; }
  .profit.redline { background: #f4f4f5; color: #999; }
  table.info { width: 100%; border-collapse: collapse; margin-bottom: 10px; }
  table.info td { font-size: 13px; padding: 3px 0; }
  table.info td:first-child { color: #888; width: 86px; }
  table.info td:last-child { font-weight: 600; }
  .btn-row { display: flex; gap: 8px; margin-bottom: 8px; }
  .btn { flex: 1; display: block; text-align: center; background: #1f4e79; color: #fff; text-decoration: none;
         border-radius: 10px; padding: 10px; font-size: 13px; font-weight: 600; }
  .btn.copy { background: #e8f0fb; color: #1f4e79; border: none; font-family: inherit; cursor: pointer; }
  .btn.dewu { background: #fff4e0; color: #b26a00; }
  .hint { font-size: 11px; color: #999; text-align: center; margin: -2px 0 6px; }
  details { margin-bottom: 6px; }
  summary { font-size: 13px; font-weight: 600; color: #1f4e79; cursor: pointer; padding: 4px 0; }
  .tut { font-size: 13px; line-height: 1.7; color: #444; background: #f7f9fb; border-radius: 8px; padding: 8px 10px; }
  .cal { font-size: 12px; color: #666; line-height: 1.6; border-top: 1px dashed #e3e6ea; padding-top: 8px; margin-top: 4px; }
  .exp { font-size: 12px; color: #1e7e45; background: #e6f6ec; line-height: 1.6; border-radius: 8px; padding: 8px 10px; margin-bottom: 8px; }
  .note { font-size: 12px; color: #b26a00; line-height: 1.6; margin-top: 6px; }
  .livebox { background: #fff; border-radius: 10px; padding: 12px; }
  .liveintro { font-size: 12px; color: #666; line-height: 1.6; margin-bottom: 8px; }
  .chips { display: flex; flex-direction: column; gap: 6px; }
  .chip { display: block; background: #f7f9fb; border: 1px solid #e3e6ea; border-radius: 8px;
          padding: 9px 10px; color: #1f4e79; text-decoration: none; font-size: 13px; }
  footer { font-size: 11px; color: #999; text-align: center; line-height: 1.7; margin-top: 18px; }
</style>
</head>
<body>
<div class="wrap">
  <header>
    <h1>📦 得物搬砖 · 差异化线报</h1>
    <p>数据更新于 __STAMP__ ｜ 共 __N__ 条 ｜ ⭐冷门优选 __COLD__ 条 ｜ 净利≥50 __OK__ 条<br>稀缺分 = 净利×时效×供需×类目×(1−曝光)，冷门优选排最前。下单前请核对到手价 &amp; 得物实时卖价。</p>
  </header>
  <div class="tabs">
    <button class="tab on" id="tabbtn-main" onclick="showTab('main')">📋 今日线报</button>
    <button class="tab" id="tabbtn-live" onclick="showTab('live')">🔴 实时线报</button>
  </div>
  <div class="stat">
    <div><div class="n">__COLD__</div><div class="t">冷门优选</div></div>
    <div><div class="n">__OK__</div><div class="t">净利≥50</div></div>
    <div><div class="n">__N__</div><div class="t">候选总数</div></div>
  </div>
  <div id="tab-main">
  __CARDS__
  </div>
  <div id="tab-live" style="display:none">
    <div class="livebox">
      <div class="liveintro">🔴 得物官方“门道商机”直达入口（点开看得物实时数据，每次更新换成最新商机）</div>
      __LIVE_LINKS__
      <div class="liveintro" style="margin-top:10px;padding-top:10px;border-top:1px dashed #e3e6ea;">❗ 全部是得物可售品类（门道商机原始数据）。⭐冷门优选优先看，普通条目标“核对费率”。</div>
    </div>
  </div>
  <footer>数据来源：门道商机（得物官方搬砖工具）自动抓取 ｜ 仅为信息整理，下单前自行核实<br>© 得物搬砖工具箱 · __STAMP__</footer>
</div>
<script>
function showTab(name){
  document.getElementById("tab-main").style.display = (name === "main") ? "" : "none";
  document.getElementById("tab-live").style.display = (name === "live") ? "" : "none";
  document.getElementById("tabbtn-main").className = "tab" + (name === "main" ? " on" : "");
  document.getElementById("tabbtn-live").className = "tab" + (name === "live" ? " on" : "");
}
function copyDewu(btn, kw){
  var done = function(){ btn.textContent = "✅ 已复制货号"; setTimeout(function(){ btn.textContent = "🔍 得物查价"; }, 2600); };
  if (navigator.clipboard && navigator.clipboard.writeText) {
    navigator.clipboard.writeText(kw).then(done).catch(function(){ fb(kw, done); });
  } else { fb(kw, done); }
}
function copyLink(btn, href){
  var done = function(){ btn.textContent = "✅ 已复制"; setTimeout(function(){ btn.textContent = "📋 复制链接"; }, 1600); };
  if (navigator.clipboard && navigator.clipboard.writeText) {
    navigator.clipboard.writeText(href).then(done).catch(function(){ fb(href, done); });
  } else { fb(href, done); }
}
function fb(t, done){
  var ta = document.createElement("textarea");
  ta.value = t; ta.style.position = "fixed"; ta.style.opacity = "0";
  document.body.appendChild(ta); ta.select();
  try { document.execCommand("copy"); done(); } catch(e) {}
  document.body.removeChild(ta);
}
</script>
</body>
</html>"""


if __name__ == "__main__":
    main()
